import os
from pathlib import Path
from moviepy.editor import VideoFileClip
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import openpyxl as px
from datetime import timedelta

VIDEO_EXTENSIONS = [".mp4", ".mkv", ".avi"]

# Function to calculate the total and average durations of the videos
def calculate_durations(videos):
    times = []
    for video in videos:
        try:
            clip = VideoFileClip(video)
            times.append(clip.duration / 60)  # Duration in minutes
            clip.close()
        except Exception as e:
            print(f"Error processing {video}: {e}")
    return times

# Function to format time into HH:MM:SS
def format_time(time_min):
    if pd.isna(time_min):
        return None
    total_seconds = int(time_min * 60)
    return str(timedelta(seconds=total_seconds))

# Main function to map videos and generate the report
def map_videos(root_path, exclude_paths=[], level=0):
    results = []
    exclude_paths = [os.path.abspath(p.strip()) for p in exclude_paths if p.strip()]

    for root, dirs, files in os.walk(root_path):
        # Ignore excluded paths
        if any(os.path.commonpath([root, ex]) == ex for ex in exclude_paths):
            continue

        current_level = root[len(root_path):].count(os.sep)
        if current_level >= level:
            videos = [os.path.join(root, f) for f in files if f.endswith(tuple(VIDEO_EXTENSIONS))]
            durations = calculate_durations(videos)

            for i, video in enumerate(videos):
                results.append({
                    "Folder Name": os.path.relpath(root, root_path).replace("\\", "/"),
                    "File Name": os.path.basename(video),
                    "Video Duration": durations[i] if i < len(durations) else None
                })

    df = pd.DataFrame(results)

    # Calculating aggregates by folder
    if not df.empty:
        df['Total Time in Folder'] = df.groupby('Folder Name')['Video Duration'].transform('sum')
        df['Average Time in Folder'] = df.groupby('Folder Name')['Video Duration'].transform('mean')

    # Total time and average time for all videos (calculated from the entire dataset)
    total_time = df['Video Duration'].sum()
    avg_time = df['Video Duration'].mean()

    # Add 'Total Average Time' and 'Total Time' columns for the final row (in merged cells)
    df["Total Average Time"] = avg_time
    df["Total Time"] = total_time

    # Format all time-related columns to HH:MM:SS
    time_columns = ["Video Duration", "Total Time in Folder", "Average Time in Folder", "Total Average Time", "Total Time"]
    for col in time_columns:
        df[col] = df[col].apply(format_time)

    return df, total_time, avg_time

# Function to verify if all cells in a column within a range have the same value
def verify_uniform_values(ws, col, start_row, end_row):
    values = [ws.cell(row=i, column=col).value for i in range(start_row, end_row + 1)]
    if all(value == values[0] for value in values):
        return True
    else:
        raise ValueError(f"Values in column {col} from row {start_row} to {end_row} are not uniform.")

# Function to save to Excel with merged cells and additional formatting
def save_with_merged_cells(df, output_path):
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Video Report"

    # Add data to Excel
    start_row, start_col = 2, 2  # Set table start at B2
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = "0070C0"
    for cell in ws[start_row]:  # Header row
        cell.font = header_font
        cell.fill = px.styles.PatternFill(start_color=header_fill, end_color=header_fill, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=start_col, max_col=start_col + len(df.columns) - 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Merge cells in the "Folder Name" column and propagate the same merge range to columns start_col+3 and start_col+4
    col_to_merge = start_col
    start_merge_row = None
    last_value = None
    for i in range(start_row + 1, ws.max_row + 1):  # Skip header
        current_value = ws.cell(row=i, column=col_to_merge).value
        if current_value == last_value:
            if start_merge_row is None:
                start_merge_row = i - 1
        else:
            if start_merge_row is not None:
                # Validate uniformity for start_col+3 and start_col+4
                verify_uniform_values(ws, col_to_merge + 3, start_merge_row, i - 1)
                verify_uniform_values(ws, col_to_merge + 4, start_merge_row, i - 1)
                
                # Merge cells in the base column and propagate to the other columns
                ws.merge_cells(start_row=start_merge_row, start_column=col_to_merge, end_row=i - 1, end_column=col_to_merge)
                ws.merge_cells(start_row=start_merge_row, start_column=col_to_merge + 3, end_row=i - 1, end_column=col_to_merge + 3)
                ws.merge_cells(start_row=start_merge_row, start_column=col_to_merge + 4, end_row=i - 1, end_column=col_to_merge + 4)
                start_merge_row = None
        last_value = current_value

    if start_merge_row is not None:
        # Validate uniformity for start_col+3 and start_col+4 for the last range
        verify_uniform_values(ws, col_to_merge + 3, start_merge_row, ws.max_row)
        verify_uniform_values(ws, col_to_merge + 4, start_merge_row, ws.max_row)
        
        # Merge remaining rows at the end of the sheet
        ws.merge_cells(start_row=start_merge_row, start_column=col_to_merge, end_row=ws.max_row, end_column=col_to_merge)
        ws.merge_cells(start_row=start_merge_row, start_column=col_to_merge + 3, end_row=ws.max_row, end_column=col_to_merge + 3)
        ws.merge_cells(start_row=start_merge_row, start_column=col_to_merge + 4, end_row=ws.max_row, end_column=col_to_merge + 4)

    # Validate uniformity and merge all rows in columns start_col+5 and start_col+6
    for col in [start_col + 5, start_col + 6]:
        verify_uniform_values(ws, col, start_row + 1, ws.max_row)
        ws.merge_cells(start_row=start_row + 1, start_column=col, end_row=ws.max_row, end_column=col)

    # Center all cells horizontally and vertically
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=start_col, max_col=(start_col + len(df.columns) - 1)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Fill all cells outside the table with white color
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Calculate the maximum range for filling
    table_end_row = start_row + len(df)  # End of the table in rows
    table_end_col = start_col + len(df.columns) - 1  # End of the table in columns

    # Adjust to cover an area beyond the table
    max_row = table_end_row + 100  # 100 extra rows
    max_col = table_end_col + 100  # 100 extra columns

    # Iterate over all cells in the range
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            # Exclude cells within the table range
            if not (start_row <= cell.row <= table_end_row and start_col <= cell.column <= table_end_col):
                cell.fill = white_fill  # Fill with white

    # Auto adjust column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter
        for cell in col:
            try:
                if cell.value:  # Only adjust if there's content
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add padding
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save the Excel file
    wb.save(output_path)

# User inputs
root_path = input("Enter the root folder path: ")
level = int(input("Enter the level of subfolders you want to include (0 for root, 1 for subfolders, etc.): ") or 0)
exclude_paths = input("Enter the paths of folders to exclude (separated by commas): ").split(',')

# Map all videos from the input foler
df_result, total_time, avg_time = map_videos(root_path, exclude_paths, level)

# Display the result
print("\nResults Report:")
print(df_result)

# Display total and average times in HH:MM:SS format
from datetime import timedelta

total_time_formatted = str(timedelta(seconds=int(total_time * 60)))
avg_time_formatted = str(timedelta(seconds=int(avg_time * 60)))

print(f"\nTotal time for all videos: {total_time_formatted}")
print(f"Average time for all videos: {avg_time_formatted}")

# Export to Excel with merged cells
create_report = (input("\nDo you want to create the report in Excel format? (y): ").strip().lower())

# Ask the user if they want to create the report
if create_report == 'y':
    output_path = os.path.join(root_path, "video_report.xlsx")
    save_with_merged_cells(df_result, output_path)
    print(f"Report saved to: {output_path}")
